from ..models.db_controller import DBController
from ..utils.styles import Colors
import os
import openpyxl
from openpyxl.styles import Font, Border
from openpyxl.styles import PatternFill, Alignment, Side
from datetime import datetime

class exportToXlsx():
    def __init__(self):
        self.db = DBController()

        # Definir bordas espessas e finas
        self.thick_border = Side(border_style="thick", color="000000")
        self.thin_border = Side(border_style="thin", color="000000")

    def export_sale_to_excel(self, date, output_file):

        sale = self.get_sale_from_db(date)
        
        if not sale:
            print(f"Nenhuma venda encontrada para a data {date}.")
            self.db.disconnect()
            return
        
        sale_id, sell_date = sale

        items = self.get_itens_from_db(sale_id)
        workbook = self.get_workbook(output_file)
        sheet_title = self.make_sheet_title(sell_date)
        sheet = workbook.create_sheet(title=sheet_title)
        headers = self.create_header(sheet, sale_id, sell_date)
        self.add_produts_to_worksheet(sheet, items)
        self.add_total_sell_to_worksheet(sheet, items)
        self.styling_table(sheet, headers)

        # Salvar o arquivo
        workbook.save(output_file)
        print(f"Arquivo salvo como {output_file}")

    def make_sheet_title(self, sell_date):
        # Converter string para datetime
        sell_date_obj = datetime.strptime(sell_date, "%Y-%m-%d %H:%M:%S")
        sheet_title = f"Venda {sell_date_obj.hour}h{sell_date_obj.minute}m"
        return sheet_title

    def add_total_sell_to_worksheet(self, sheet, items):
        # Adicionar total da venda
        sheet.append([])
        total_venda = sum(row[4] for row in items)
        total_row = ["Total da Venda:", "", "", "", f"R$ {total_venda:,.2f}"]
        sheet.append(total_row)

    def add_produts_to_worksheet(self, sheet, items):
        # Adicionar os produtos na planilha
        for item in items:
            product_name, barcode, quantity, price, total_price = item
            # Formatando o preço em reais
            price_str = f"R$ {price:,.2f}"
            total_price_str = f"R$ {total_price:,.2f}"

            sheet.append([product_name, barcode, quantity, price_str, total_price_str])

    def get_workbook(self, output_file):
        # Verifica se o caminho do arquivo existe
        if os.path.exists(output_file):
            workbook = openpyxl.load_workbook(output_file)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
        return workbook

    def create_header(self, sheet, sale_id, sell_date):
        # Mesclar células das duas primeiras linhas
        sheet.merge_cells('A1:E1')
        sheet.merge_cells('A2:E2')

        # Adicionar Título (Venda ID) e Subtítulo (Data da venda)
        title_cell = sheet['A1']
        title_cell.value = f"Venda {sale_id}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.fill = PatternFill(start_color=Colors.violetButton.replace("#", ""), fill_type="solid")
        title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')


        subtitle_cell = sheet['A2']
        subtitle_cell.value = f"Data: {sell_date}"
        subtitle_cell.font = Font(size=12, italic=True, bold=True)
        subtitle_cell.fill = PatternFill(start_color=Colors.violetButton.replace("#", ""), fill_type="solid")
        subtitle_cell.alignment = Alignment(horizontal='center')

        headers = ["Produto", "Código de Barras", "Quantidade", "Preço Unitário", "Subtotal"]
        sheet.append(headers)

        # Retorna cabeçalhos da tabela
        return headers
    
    def styling_table(self, sheet, headers):
        # Estilizar os cabeçalhos
        for col_num, header in enumerate(headers, start=1):  
            cell = sheet.cell(row=3, column=col_num)  
            cell.font = Font(size=11, bold=True)  # Negrito e tamanho da fonte  
            cell.fill = PatternFill(start_color=Colors.xlsxHeader.replace("#", ""), fill_type="solid")  # Fundo roxo (violetButton)  
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Centralização  
            cell.border = Border(bottom=self.thin_border)
        
        # Ajusta largura das colunas automaticamente
        for col_cells in sheet.columns:
            col_cells = [cell for cell in col_cells if not isinstance(cell, openpyxl.cell.MergedCell)]  # Filtra células mescladas

            if col_cells:  # Garante que há células válidas
                col_letter = col_cells[0].column_letter
                max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
                sheet.column_dimensions[col_letter].width = max_length + 1

        # Centralizando a coluna 'Código de Barras'
        for row in sheet.iter_rows(min_row=4, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        # Centralizar a coluna 'Quantidade'
        for row in sheet.iter_rows(min_row=4, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
        
        # Alinhando a direita a coluna 'Preço unitário'
        for row in sheet.iter_rows(min_row=4, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal='right')
        
        # Alinhando a direita a coluna 'Preço unitário'
        for row in sheet.iter_rows(min_row=4, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal='right')

        # Definindo bordas externas
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical
                )
                cell.border = Border(
                    left=self.thick_border if cell.column == 1 else None,   # Borda esquerda espessa
                    right=self.thick_border if cell.column == 5 else None,  # Borda direita espessa
                    top=self.thick_border if cell.row == 1 else None,       # Borda superior espessa
                    bottom=self.thick_border if cell.row == sheet.max_row else None,  # Borda inferior espessa
                )

        # Definindo bordas do header
        for row in sheet.iter_rows(min_row=3, max_row=3, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical
                )
                cell.border = Border(
                    bottom=self.thin_border,  # Borda inferior espessa
                    left=self.thick_border if cell.column == 1 else self.thin_border,
                    right=self.thick_border if cell.column == 5 else self.thin_border,
                    top=self.thick_border,
                )

        # Aplica cor de violetBackground em toda a tabela
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.fill = PatternFill(start_color=Colors.violetBackground.replace("#", ""), fill_type="solid")

        # Estiliza a tupla total_sells
        sheet["A" + str(sheet.max_row)].font = Font(bold=True)
        sheet["E" + str(sheet.max_row)].font = Font(bold=True)
        sheet["E" + str(sheet.max_row)].alignment = Alignment(horizontal='right')

    def get_sale_from_db(self, date):
        # Busca a venda do banco de dados
        self.db.connect()
        self.db.cursor.execute("""
            SELECT id, sell_date 
            FROM tb_sells 
            WHERE DATE(sell_date) = ?
            ORDER BY sell_date DESC
        """, (date,))
        sale = self.db.cursor.fetchone()
        self.db.disconnect()
        return sale
    
    def get_itens_from_db(self, sale_id):
        # Busca itens do banco de dados
        self.db.connect()
        self.db.cursor.execute("""
            SELECT s.product_name, s.barcode, SUM(si.quantity), s.price, SUM(si.total_price) 
            FROM tb_selled_items si
            JOIN tb_storage s ON si.product_id = s.id
            WHERE si.sell_id = ? AND s.is_active = 1
            GROUP BY s.product_name, s.barcode, s.price
            ORDER BY SUM(si.quantity) DESC
        """, (sale_id,))
        items = self.db.cursor.fetchall()
        self.db.disconnect()
        return items
        


